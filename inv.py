import streamlit as st 
import pandas as pd
from datetime import datetime 
from fpdf import FPDF 
from streamlit_gsheets import GSheetsConnection
import io

# ------- CONFIGURACIÓN DE PÁGINA ---------
st.set_page_config(
    page_title="Control de Inventario GCM", 
    page_icon="📦",
    layout="wide",
    initial_sidebar_state="collapsed"
) 

# 🎨 CSS PERSONALIZADO PARA TÍTULO Y TABLAS
st.markdown("""
    <style>
    /* Ocultar completamente la barra superior (HUD) y cualquier rastro de sidebar de Streamlit */
    header[data-testid="stHeader"], [data-testid="stSidebar"] {
        display: none !important;
    }
    
    /* Centrar y estilizar el título principal a lo largo */
    .titulo-principal {
        text-align: center;
        font-weight: 700;
        font-size: 2.5rem;
        margin-bottom: 0px;
        padding-top: 10px;
    }

    /* Ajustes para las tablas */
    [data-testid="stDataFrame"] div[data-testid="stTable"] td,
    .stDataFrame td, div[data-baseweb="datatable"] div {
        white-space: normal !important;
        word-wrap: break-word !important;
        height: auto !important;
    }
    </style>
""", unsafe_allow_html=True)

# 🔐 USUARIOS AUTORIZADOS
USUARIOS = {
    "zaid": "2010",
    "jdiaz": "1978",
    "ccdiazj": "1974",
    "gael": "2003",
    "monica": "2026",
    "sergio": "sergio2026*"
}

# 🔐 FUNCIÓN LOGIN
def login():
    st.title("🔐 Iniciar sesión")
    usuario = st.text_input("Usuario")
    password = st.text_input("Contraseña", type="password")

    if st.button("Entrar"):
        if usuario in USUARIOS and USUARIOS[usuario] == password:
            st.session_state["autenticado"] = True
            st.session_state["usuario"] = usuario
            st.rerun()
        else:
            st.error("Usuario o contraseña incorrectos")

# 🔒 CONTROL DE SESIÓN
if "autenticado" not in st.session_state:
    st.session_state["autenticado"] = False

if not st.session_state["autenticado"]:
    login()
    st.stop()

# -------- CABECERA PRINCIPAL (Usuario, Título y Botones arriba con texto) ----------
col_top1, col_top2, col_top3 = st.columns([2, 6, 2])

with col_top1:
    st.write(f"👤 **{st.session_state['usuario']}**")
    c_btn1, c_btn2 = st.columns(2)
    with c_btn1:
        if st.button("🔄 Actualizar"):
            st.cache_data.clear()
            st.session_state["menu_opcion"] = "Registrar Producto"  # Restablece el menú al actualizar
            st.rerun()
    with c_btn2:
        if st.button("🚪 Salir"):
            st.session_state["autenticado"] = False
            st.rerun()

with col_top2:
    st.markdown("<h1 class='titulo-principal'>Control de Inventario GCM</h1>", unsafe_allow_html=True)

with col_top3:
    try:
        st.image("LOGOGCM.jpeg", width=110)
    except Exception:
        pass

st.markdown("---")

# -------- MENÚ DE OPCIONES PRINCIPAL ----------
menu = st.selectbox(
    "Selecciona una opción", 
    ["Registrar Producto", "Registrar Movimiento", "Ver Inventario", "Historial de Movimientos", "Solicitar Pedidos"],
    key="menu_opcion"
)

st.markdown("---")

# -------Conexión a Google Sheets--------- 
conn = st.connection("gsheets", type=GSheetsConnection)

def cargar_inventario_dataframe(): 
    try:
        df = conn.read(worksheet="Productos", ttl=0)
        if df is None or df.empty:
            return pd.DataFrame(columns=["PRODUCTO", "CANTIDAD TOTAL", "CANTIDAD ABIERTOS", "ALMACEN", "FECHA INGRESO"])
        
        df.columns = df.columns.astype(str).str.strip().str.upper()
        
        if "CANTIDAD_TOTAL" in df.columns and "CANTIDAD TOTAL" not in df.columns:
            df.rename(columns={"CANTIDAD_TOTAL": "CANTIDAD TOTAL"}, inplace=True)
        if "CANTIDAD_ABIERTOS" in df.columns and "CANTIDAD ABIERTOS" not in df.columns:
            df.rename(columns={"CANTIDAD_ABIERTOS": "CANTIDAD ABIERTOS"}, inplace=True)
        if "FECHA_INGRESO" in df.columns and "FECHA INGRESO" not in df.columns:
            df.rename(columns={"FECHA_INGRESO": "FECHA INGRESO"}, inplace=True)
            
        if "PRODUCTO" not in df.columns:
            return pd.DataFrame(columns=["PRODUCTO", "CANTIDAD TOTAL", "CANTIDAD ABIERTOS", "ALMACEN", "FECHA INGRESO"])
        
        if "CANTIDAD TOTAL" not in df.columns:
            df["CANTIDAD TOTAL"] = 0
        if "CANTIDAD ABIERTOS" not in df.columns:
            df["CANTIDAD ABIERTOS"] = 0
        if "ALMACEN" not in df.columns:
            df["ALMACEN"] = 1
        if "FECHA INGRESO" not in df.columns:
            df["FECHA INGRESO"] = ""
            
        df = df.dropna(subset=["PRODUCTO"])
        df = df[df["PRODUCTO"].astype(str).str.strip() != ""]
        
        return df
    except Exception:
        return pd.DataFrame(columns=["PRODUCTO", "CANTIDAD TOTAL", "CANTIDAD ABIERTOS", "ALMACEN", "FECHA INGRESO"])

def cargar_historial_dataframe():
    try:
        df = conn.read(worksheet="Historial", ttl=0)
        if df is None or df.empty:
            return pd.DataFrame(columns=["PRODUCTO", "TIPO", "CANTIDAD", "ABIERTOS", "FECHA", "DESCRIPCION"])
        df.columns = df.columns.astype(str).str.strip().str.upper()
        return df
    except Exception:
        return pd.DataFrame(columns=["PRODUCTO", "TIPO", "CANTIDAD", "ABIERTOS", "FECHA", "DESCRIPCION"])

def limpiar_entero(valor):
    try:
        return int(float(valor))
    except (ValueError, TypeError):
        return 0

def limpiar_texto_pdf(texto):
    return str(texto).encode('latin-1', 'replace').decode('latin-1')

def registrar_en_historial(producto, tipo, cantidad, abiertos, fecha, descripcion=""):
    try:
        df_historial = conn.read(worksheet="Historial", ttl=0)
        if df_historial is None or df_historial.empty:
            df_historial = pd.DataFrame(columns=["PRODUCTO", "TIPO", "CANTIDAD", "ABIERTOS", "FECHA", "DESCRIPCION"])
    except Exception:
        df_historial = pd.DataFrame(columns=["PRODUCTO", "TIPO", "CANTIDAD", "ABIERTOS", "FECHA", "DESCRIPCION"])
    
    nuevo_registro = pd.DataFrame([{
        "PRODUCTO": producto,
        "TIPO": tipo,
        "CANTIDAD": limpiar_entero(cantidad),
        "ABIERTOS": limpiar_entero(abiertos),
        "FECHA": fecha,
        "DESCRIPCION": descripcion
    }])
    
    df_historial_actualizado = pd.concat([df_historial, nuevo_registro], ignore_index=True)
    conn.update(worksheet="Historial", data=df_historial_actualizado.astype(str))

def registrar_en_hoja_long(usuario, accion, producto, cantidad, fecha):
    try:
        df_long = conn.read(worksheet="long", ttl=0)
        if df_long is None or df_long.empty:
            df_long = pd.DataFrame(columns=["USUARIO", "ACCION", "PRODUCTO", "CANTIDAD", "ABIERTOS", "FECHA"])
    except Exception:
        df_long = pd.DataFrame(columns=["USUARIO", "ACCION", "PRODUCTO", "CANTIDAD", "ABIERTOS", "FECHA"])
    
    nuevo_registro = pd.DataFrame([{
        "USUARIO": usuario,
        "ACCION": accion,
        "PRODUCTO": producto,
        "CANTIDAD": limpiar_entero(cantidad),
        "ABIERTOS": limpiar_entero(cantidad),
        "FECHA": fecha
    }])
    
    df_long_actualizado = pd.concat([df_long, nuevo_registro], ignore_index=True)
    conn.update(worksheet="long", data=df_long_actualizado.astype(str))

# ---------REGISTRO DE NUEVO PRODUCTO----------- 
if menu == "Registrar Producto": 
    st.subheader("Registrar Nuevo Producto") 

    if "nombre" not in st.session_state: st.session_state.nombre = ""
    if "cantidad_total" not in st.session_state: st.session_state.cantidad_total = 0
    if "cantidad_abiertos" not in st.session_state: st.session_state.cantidad_abiertos = 0
    if "almacen" not in st.session_state: st.session_state.almacen = 1
    if "fecha_ingreso" not in st.session_state: st.session_state.fecha_ingreso = datetime.now()

    def procesar_y_limpiar():
        nom = st.session_state.nombre
        tot = st.session_state.cantidad_total
        ab = st.session_state.cantidad_abiertos
        alm = st.session_state.almacen
        fecha_str = st.session_state.fecha_ingreso.strftime('%Y-%m-%d')
        
        if nom.strip():
            if ab > tot:
                st.error("❌ Los abiertos no pueden ser mayores a la cantidad total.")
            else:
                df_productos = cargar_inventario_dataframe()
                nombre_mayus = nom.strip().upper()
                if not df_productos.empty and "PRODUCTO" in df_productos.columns: 
                    if nombre_mayus in df_productos["PRODUCTO"].astype(str).str.strip().str.upper().values: 
                        st.warning("El producto ya existe en el inventario.") 
                        return
                
                cant_total_limpia = limpiar_entero(tot)
                cant_abiertos_limpia = limpiar_entero(ab)

                nuevo_prod_df = pd.DataFrame([{
                    "PRODUCTO": nombre_mayus,
                    "CANTIDAD TOTAL": cant_total_limpia,
                    "CANTIDAD ABIERTOS": cant_abiertos_limpia,
                    "ALMACEN": alm,
                    "FECHA INGRESO": fecha_str
                }])
                
                df_actualizado = pd.concat([df_productos, nuevo_prod_df], ignore_index=True)
                conn.update(worksheet="Productos", data=df_actualizado.astype(str))
                
                registrar_en_historial(nombre_mayus, "Entrada (Inicial)", cant_total_limpia, cant_abiertos_limpia, fecha_str, f"Inicial - Abiertos: {cant_abiertos_limpia}")
                usuario_actual = st.session_state.get("usuario", "sistema")
                registrar_en_hoja_long(usuario_actual, "Entrada (Inicial)", nombre_mayus, cant_total_limpia, fecha_str)
                
                st.session_state.nombre = ""
                st.session_state.cantidad_total = 0
                st.session_state.cantidad_abiertos = 0
                st.session_state.almacen = 1
                st.toast("✅ Producto registrado correctamente.")
        else:
            st.warning("Completa todos los campos.")

    st.text_input("Nombre del producto", key="nombre") 
    st.number_input("Cantidad total inicial", min_value=0, step=1, key="cantidad_total") 
    st.number_input("¿Cuántos de estos son abiertos y en buen estado?", min_value=0, max_value=int(st.session_state.cantidad_total), step=1, key="cantidad_abiertos")
    st.number_input("Almacén", min_value=1.0, max_value=99.0, step=0.1, key="almacen")
    st.date_input("Fecha de ingreso", key="fecha_ingreso")
    
    st.button("Registrar Producto", on_click=procesar_y_limpiar)

# --------MOVIMIENTOS--------
elif menu == "Registrar Movimiento": 
    st.subheader("Registrar Movimiento de Stock") 
    df_productos = cargar_inventario_dataframe() 
    
    if df_productos.empty: 
        st.info("No hay productos registrados en Google Sheets o la tabla está vacía.") 
    else: 
        productos = sorted(df_productos["PRODUCTO"].astype(str).tolist()) 
        nombre = st.selectbox("Producto", productos, index=None, placeholder="Selecciona un producto...") 
        
        tipo_mov = st.radio("Tipo", ["Entrada", "Salida"], horizontal=True)
            
        cantidad = st.number_input("Cantidad total a mover", min_value=1, step=1) 
        abiertos_afectados = st.number_input("¿Cuántos de estos son abiertos y en buen estado?", min_value=0, max_value=int(cantidad), step=1)
        
        descripcion = st.text_input("Descripción (Opcional)", placeholder="Escribe un detalle o motivo opcional...")
        
        if st.button("Registrar Movimiento"): 
            if not nombre:
                st.warning("Por favor, selecciona un producto.")
            else:
                idx = df_productos[df_productos["PRODUCTO"] == nombre].index[0]
                cant_total_actual = limpiar_entero(df_productos.loc[idx, "CANTIDAD TOTAL"])
                cant_abiertos_actual = limpiar_entero(df_productos.loc[idx, "CANTIDAD ABIERTOS"])
                cant_mov = limpiar_entero(cantidad)
                abiertos_mov = limpiar_entero(abiertos_afectados)
                    
                if tipo_mov == "Salida":
                    if cant_mov > cant_total_actual or abiertos_mov > cant_abiertos_actual: 
                        st.warning("No hay suficiente stock total o abiertos para realizar la salida.") 
                    else:
                        nueva_total = cant_total_actual - cant_mov
                        nuevo_abiertos = cant_abiertos_actual - abiertos_mov
                        df_productos.loc[idx, "CANTIDAD TOTAL"] = nueva_total
                        df_productos.loc[idx, "CANTIDAD ABIERTOS"] = nuevo_abiertos
                        conn.update(worksheet="Productos", data=df_productos.astype(str))
                        
                        fecha_hoy = datetime.now().strftime('%Y-%m-%d')
                        registrar_en_historial(nombre, tipo_mov, cant_mov, abiertos_mov, fecha_hoy, f"{descripcion} (Abiertos: {abiertos_mov})")
                        usuario_actual = st.session_state.get("usuario", "sistema")
                        registrar_en_hoja_long(usuario_actual, tipo_mov, nombre, cant_mov, fecha_hoy)
                        st.success("Movimiento de salida registrado con éxito.")
                        st.rerun()
                else: 
                    nueva_total = cant_total_actual + cant_mov
                    nuevo_abiertos = cant_abiertos_actual + abiertos_mov
                    df_productos.loc[idx, "CANTIDAD TOTAL"] = nueva_total
                    df_productos.loc[idx, "CANTIDAD ABIERTOS"] = nuevo_abiertos
                    conn.update(worksheet="Productos", data=df_productos.astype(str))
                    
                    fecha_hoy = datetime.now().strftime('%Y-%m-%d')
                    registrar_en_historial(nombre, tipo_mov, cant_mov, abiertos_mov, fecha_hoy, f"{descripcion} (Abiertos: {abiertos_mov})")
                    usuario_actual = st.session_state.get("usuario", "sistema")
                    registrar_en_hoja_long(usuario_actual, tipo_mov, nombre, cant_mov, fecha_hoy)
                    st.success("Movimiento de entrada registrado con éxito.")
                    st.rerun()

# --------VER INVENTARIO--------
elif menu == "Ver Inventario": 
    st.subheader("Reporte General de Inventario") 
    st.caption("Puedes corregir nombre, cantidades, almacén y fecha de ingreso.")

    df_productos = cargar_inventario_dataframe()
    
    if df_productos.empty: 
        st.info("No hay productos registrados.") 
    else: 
        lista = [] 
        for _, row in df_productos.iterrows(): 
            nombre = str(row["PRODUCTO"])
            fecha_ing = str(row.get("FECHA INGRESO", "-"))
            
            lista.append({ 
                "Producto Original": nombre,
                "PRODUCTO": nombre, 
                "CANTIDAD TOTAL": limpiar_entero(row["CANTIDAD TOTAL"]), 
                "CANTIDAD ABIERTOS": limpiar_entero(row["CANTIDAD ABIERTOS"]), 
                "ALMACÉN": limpiar_entero(row["ALMACEN"]),
                "FECHA INGRESO": fecha_ing 
            }) 
        
        tabla = st.data_editor(
            lista,
            column_config={
                "Producto Original": None,
                "PRODUCTO": st.column_config.TextColumn(),
                "CANTIDAD TOTAL": st.column_config.NumberColumn(alignment="center", format="%d"),
                "CANTIDAD ABIERTOS": st.column_config.NumberColumn(alignment="center", format="%d"),
                "ALMACÉN": st.column_config.NumberColumn(alignment="center", min_value=1.0, max_value=99.0, step=0.1, format="%.1f"),
                "FECHA INGRESO": st.column_config.TextColumn(alignment="center")
            },
            hide_index=True,
            use_container_width=True
        )

        if tabla != lista:
            nombres = set()
            error = False

            for fila in tabla:
                nuevo_nombre = fila["PRODUCTO"].strip().upper()
                if nuevo_nombre in nombres:
                    st.error(f"❌ El producto '{nuevo_nombre}' está duplicado.")
                    error = True
                    break
                if fila["CANTIDAD ABIERTOS"] > fila["CANTIDAD TOTAL"]:
                    st.error(f"❌ Los abiertos de '{nuevo_nombre}' no pueden superar al total.")
                    error = True
                    break
                nombres.add(nuevo_nombre)

            if not error:
                nuevos_productos = []
                for fila in tabla:
                    nuevo_nombre = fila["PRODUCTO"].strip().upper()
                    nueva_total = limpiar_entero(fila["CANTIDAD TOTAL"])
                    nuevo_abiertos = limpiar_entero(fila["CANTIDAD ABIERTOS"])
                    nueva_fecha_ing = fila["FECHA INGRESO"]

                    nuevos_productos.append({
                        "PRODUCTO": nuevo_nombre,
                        "CANTIDAD TOTAL": nueva_total,
                        "CANTIDAD ABIERTOS": nuevo_abiertos,
                        "ALMACEN": fila["ALMACÉN"],
                        "FECHA INGRESO": nueva_fecha_ing
                    })

                df_nuevos_prod = pd.DataFrame(nuevos_productos)
                conn.update(worksheet="Productos", data=df_nuevos_prod.astype(str))

                st.success("✅ Cambios guardados correctamente en Google Sheets.")
                st.rerun()

        class PDFReporte(FPDF):
            def __init__(self):
                super().__init__()
                self.col_widths = [65, 25, 25, 25, 50]

            def header(self):
                if self.page_no() == 1:
                    self.set_font("Arial", 'B', 16)
                    self.cell(0, 10, "Reporte de Inventario", ln=True, align='C')
                    self.ln(5)
                
                self.set_font("Arial", 'B', 9)
                headers = ["PRODUCTO", "TOTAL", "ABIERTOS", "ALMACEN", "FECHA INGRESO"]
                for i, h in enumerate(headers):
                    self.cell(self.col_widths[i], 8, h, 1, 0, 'C')
                self.ln()

            def add_row_multi_cell(self, row_data):
                self.set_font("Arial", '', 9)
                prod_text = limpiar_texto_pdf(row_data[0])
                
                w_prod = self.col_widths[0]
                line_height = 5
                
                words = prod_text.split(' ')
                lines = []
                current_line = ""
                
                for word in words:
                    if self.get_string_width(word) > (w_prod - 4):
                        if current_line:
                            lines.append(current_line)
                            current_line = ""
                        sub_word = ""
                        for char in word:
                            if self.get_string_width(sub_word + char) <= (w_prod - 4):
                                sub_word += char
                            else:
                                lines.append(sub_word)
                                sub_word = char
                        if sub_word:
                            current_line = sub_word
                    else:
                        test_line = f"{current_line} {word}".strip()
                        if self.get_string_width(test_line) <= (w_prod - 4):
                            current_line = test_line
                        else:
                            if current_line:
                                lines.append(current_line)
                            current_line = word
                if current_line:
                    lines.append(current_line)
                if not lines:
                    lines = [""]

                num_lines = len(lines)
                row_height = max(8, num_lines * line_height)

                if self.get_y() + row_height > self.page_break_trigger:
                    self.add_page()

                x_start = self.get_x()
                y_start = self.get_y()

                self.rect(x_start, y_start, self.col_widths[0], row_height)
                self.rect(x_start + self.col_widths[0], y_start, self.col_widths[1], row_height)
                self.rect(x_start + self.col_widths[0] + self.col_widths[1], y_start, self.col_widths[2], row_height)
                self.rect(x_start + sum(self.col_widths[:3]), y_start, self.col_widths[3], row_height)
                self.rect(x_start + sum(self.col_widths[:4]), y_start, self.col_widths[4], row_height)

                self.set_xy(x_start, y_start + (row_height - (num_lines * line_height)) / 2)
                for line in lines:
                    self.cell(w_prod, line_height, line, 0, 0, 'L')
                    self.set_xy(x_start, self.get_y() + line_height)

                y_cols = y_start + (row_height - 5) / 2
                self.set_xy(x_start + self.col_widths[0], y_cols)
                self.cell(self.col_widths[1], 5, str(row_data[1]), 0, 0, 'C')

                self.set_xy(x_start + sum(self.col_widths[:2]), y_cols)
                self.cell(self.col_widths[2], 5, str(row_data[2]), 0, 0, 'C')

                self.set_xy(x_start + sum(self.col_widths[:3]), y_cols)
                self.cell(self.col_widths[3], 5, str(row_data[3]), 0, 0, 'C')

                self.set_xy(x_start + sum(self.col_widths[:4]), y_cols)
                self.cell(self.col_widths[4], 5, str(row_data[4]), 0, 0, 'C')

                self.set_xy(x_start, y_start + row_height)

        def exportar_pdf(): 
            pdf = PDFReporte() 
            pdf.add_page() 
            for _, row in df_productos.iterrows(): 
                prod_nombre = str(row["PRODUCTO"])
                tot = str(limpiar_entero(row["CANTIDAD TOTAL"]))
                ab = str(limpiar_entero(row["CANTIDAD ABIERTOS"]))
                almacen = str(row["ALMACEN"])
                fecha_ing = str(row.get("FECHA INGRESO", "-"))
                
                pdf.add_row_multi_cell([prod_nombre, tot, ab, almacen, fecha_ing])
            return bytes(pdf.output(dest='S'))

        def exportar_excel_general(df, titulo_pestana="Inventario"):
            output = io.BytesIO()
            with pd.ExcelWriter(output, engine='openpyxl') as writer:
                df.to_excel(writer, index=False, sheet_name=titulo_pestana)
                workbook = writer.book
                worksheet = writer.sheets[titulo_pestana]
                
                from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
                header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
                header_font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
                centered_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                left_alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
                
                thin_border = Border(
                    left=Side(style='thin', color='D9D9D9'), right=Side(style='thin', color='D9D9D9'),
                    top=Side(style='thin', color='D9D9D9'), bottom=Side(style='thin', color='D9D9D9')
                )

                for col_idx, col_name in enumerate(df.columns, 1):
                    cell_letter = worksheet.cell(row=1, column=col_idx).coordinate[0]
                    header_cell = worksheet[f"{cell_letter}1"]
                    header_cell.fill = header_fill
                    header_cell.font = header_font
                    header_cell.alignment = centered_alignment
                    
                    max_len = len(str(col_name))
                    for row_idx in range(2, len(df) + 2):
                        cell = worksheet[f"{cell_letter}{row_idx}"]
                        cell.border = thin_border
                        cell.font = Font(name="Arial", size=10)
                        val_str = str(cell.value or '')
                        if len(val_str) > max_len:
                            max_len = len(val_str)
                        
                        col_upper = str(col_name).upper()
                        if any(k in col_upper for k in ["CANTIDAD", "TOTAL", "ABIERTOS", "ALMACEN", "FECHA", "TIPO"]):
                            cell.alignment = centered_alignment
                        else:
                            cell.alignment = left_alignment
                    worksheet.column_dimensions[cell_letter].width = min(max(max_len + 5, 12), 50)

            output.seek(0)
            return output.getvalue()

        col_pdf, col_xlsx = st.columns(2)
        with col_pdf:
            st.download_button("📥 Descargar Inventario PDF", exportar_pdf(), "reporte_inventario.pdf", mime="application/pdf")
        with col_xlsx:
            df_excel_inv = pd.DataFrame(lista).drop(columns=["Producto Original"])
            st.download_button("📊 Descargar Inventario Excel", exportar_excel_general(df_excel_inv, "Inventario"), "reporte_inventario.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

# --------HISTORIAL--------
elif menu == "Historial de Movimientos":
    st.subheader("Historial de Movimientos")
    df_historial = cargar_historial_dataframe()

    def exportar_excel_general(df, titulo_pestana="Inventario"):
        output = io.BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name=titulo_pestana)
        output.seek(0)
        return output.getvalue()

    if df_historial.empty:
        st.info("No hay movimientos registrados.")
    else:
        st.dataframe(
            df_historial,
            column_config={
                "CANTIDAD": st.column_config.NumberColumn(alignment="center", format="%d"),
                "ABIERTOS": st.column_config.NumberColumn(alignment="center", format="%d"),
                "FECHA": st.column_config.TextColumn(alignment="center"),
                "DESCRIPCION": st.column_config.TextColumn(width="large")
            },
            hide_index=True,
            use_container_width=True
        )
        st.download_button("📊 Descargar Historial Excel", exportar_excel_general(df_historial, "Historial"), "historial_movimientos.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

# --------PEDIDOS--------
elif menu == "Solicitar Pedidos":
    st.subheader("Productos por Agotarse (Stock Total en 0 o 1)")
    df_productos = cargar_inventario_dataframe()
    bajos = []

    def exportar_excel_general(df, titulo_pestana="Inventario"):
        output = io.BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name=titulo_pestana)
        output.seek(0)
        return output.getvalue()

    if not df_productos.empty:
        for _, row in df_productos.iterrows():
            tot = limpiar_entero(row["CANTIDAD TOTAL"])
            if tot <= 1:
                bajos.append({
                    "PRODUCTO": row["PRODUCTO"],
                    "CANTIDAD TOTAL": tot,
                    "ALMACÉN": limpiar_entero(row["ALMACEN"])
                })

    if not bajos:
        st.success("✅ Todo bien de stock")
    else:
        st.error(f"⚠️ {len(bajos)} productos requieren pedido")
        df_bajos = pd.DataFrame(bajos)
        
        st.dataframe(
            df_bajos,
            column_config={
                "PRODUCTO": st.column_config.TextColumn(),
                "CANTIDAD TOTAL": st.column_config.NumberColumn(alignment="center", format="%d"),
                "ALMACÉN": st.column_config.NumberColumn(alignment="center", format="%d")
            },
            hide_index=True,
            use_container_width=True
        )

        class PDFPedidosReporte(FPDF):
            def __init__(self):
                super().__init__()
                self.col_widths = [90, 30, 30, 40]

            def header(self):
                if self.page_no() == 1:
                    self.set_font("Arial", 'B', 16)
                    self.cell(0, 10, "Reporte de Productos por Agotarse", ln=True, align='C')
                    self.ln(5)
                
                self.set_font("Arial", 'B', 10)
                headers = ["PRODUCTO", "TOTAL", "ABIERTOS", "ALMACEN"]
                for i, h in enumerate(headers):
                    self.cell(self.col_widths[i], 8, h, 1, 0, 'C')
                self.ln()

            def add_row_multi_cell(self, row_data):
                self.set_font("Arial", '', 9)
                prod_text = limpiar_texto_pdf(row_data[0])
                w_prod = self.col_widths[0]
                line_height = 5
                
                words = prod_text.split(' ')
                lines = []
                current_line = ""
                for word in words:
                    if self.get_string_width(word) > (w_prod - 4):
                        if current_line:
                            lines.append(current_line)
                            current_line = ""
                        sub_word = ""
                        for char in word:
                            if self.get_string_width(sub_word + char) <= (w_prod - 4):
                                sub_word += char
                            else:
                                lines.append(sub_word)
                                sub_word = char
                        if sub_word:
                            current_line = sub_word
                    else:
                        test_line = f"{current_line} {word}".strip()
                        if self.get_string_width(test_line) <= (w_prod - 4):
                            current_line = test_line
                        else:
                            if current_line:
                                lines.append(current_line)
                            current_line = word
                if current_line:
                    lines.append(current_line)
                if not lines:
                    lines = [""]

                num_lines = len(lines)
                row_height = max(8, num_lines * line_height)

                if self.get_y() + row_height > self.page_break_trigger:
                    self.add_page()

                x_start = self.get_x()
                y_start = self.get_y()

                self.rect(x_start, y_start, self.col_widths[0], row_height)
                self.rect(x_start + self.col_widths[0], y_start, self.col_widths[1], row_height)
                self.rect(x_start + self.col_widths[0] + self.col_widths[1], y_start, self.col_widths[2], row_height)
                self.rect(x_start + sum(self.col_widths[:3]), y_start, self.col_widths[3], row_height)

                self.set_xy(x_start, y_start + (row_height - (num_lines * line_height)) / 2)
                for line in lines:
                    self.cell(w_prod, line_height, line, 0, 0, 'L')
                    self.set_xy(x_start, self.get_y() + line_height)

                y_cols = y_start + (row_height - 5) / 2
                self.set_xy(x_start + self.col_widths[0], y_cols)
                self.cell(self.col_widths[1], 5, str(row_data[1]), 0, 0, 'C')

                self.set_xy(x_start + sum(self.col_widths[:2]), y_cols)
                self.cell(self.col_widths[2], 5, str(row_data[2]), 0, 0, 'C')

                self.set_xy(x_start + sum(self.col_widths[:3]), y_cols)
                self.cell(self.col_widths[3], 5, str(row_data[3]), 0, 0, 'C')

                self.set_xy(x_start, y_start + row_height)

        def exportar_pdf_pedidos(): 
            pdf = PDFPedidosReporte() 
            pdf.add_page() 
            for _, row in df_productos.iterrows(): 
                tot = limpiar_entero(row["CANTIDAD TOTAL"])
                if tot <= 1: 
                    prod_nombre = str(row["PRODUCTO"])
                    tot_str = str(tot)
                    ab_str = str(limpiar_entero(row["CANTIDAD ABIERTOS"]))
                    almacen_str = str(row["ALMACEN"])
                    pdf.add_row_multi_cell([prod_nombre, tot_str, ab_str, almacen_str])
            return bytes(pdf.output(dest='S'))

        col_p1, col_p2 = st.columns(2)
        with col_p1:
            st.download_button("📥 Descargar Pedidos PDF", exportar_pdf_pedidos(), "reporte_pedidos.pdf", mime="application/pdf")
        with col_p2:
            st.download_button("📊 Descargar Pedidos Excel", exportar_excel_general(df_bajos, "Pedidos"), "reporte_pedidos.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")